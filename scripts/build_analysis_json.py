#!/usr/bin/env python3
"""สร้าง analysis.json สำหรับหน้า analysis.html จากข้อมูลวิเคราะห์คลังบทข่าว 4 ปี

แหล่งข้อมูล:
  - ~/Documents/boonnews_work/รวมบทความ/สรุปข่าวรายวัด_2565-2568.xlsx
      ชีต "รายแหล่ง"        251 แหล่งข่าว × ปี 2565-2568
      ชีต "รายเดือน Top60"  60 แหล่งข่าว × 46 เดือน (ม.ค.65–ต.ค.68)
  - ~/Documents/boonnews_work/data/all_centers_profiles_v2.csv
      quadrant FB active × ส่งข่าว (คอลัมน์ active_level, news4y)

ใช้: python3 scripts/build_analysis_json.py [--out analysis.json]
ต้องมี openpyxl: pip3 install openpyxl
"""
import argparse
import csv
import json
import sys
from datetime import date
from pathlib import Path

WORK = Path.home() / 'Documents' / 'boonnews_work'
XLSX = WORK / 'รวมบทความ' / 'สรุปข่าวรายวัด_2565-2568.xlsx'
PROFILES = WORK / 'data' / 'all_centers_profiles_v2.csv'


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--out', type=Path, default=Path(__file__).resolve().parent.parent / 'analysis.json')
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        sys.exit('ต้องติดตั้ง openpyxl ก่อน: pip3 install openpyxl')
    if not XLSX.exists():
        sys.exit(f'ไม่พบ {XLSX}')
    if not PROFILES.exists():
        sys.exit(f'ไม่พบ {PROFILES}')

    wb = openpyxl.load_workbook(XLSX, read_only=True)

    # ---- รายแหล่ง: 251 แหล่ง × ปี ----
    sources = []
    ws = wb['รายแหล่ง']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        years = [int(v or 0) for v in row[4:8]]
        sources.append({
            'name': str(row[1]).strip(),
            'cat': str(row[2] or '').strip(),
            'country': str(row[3] or '').strip(),
            'y': years,
            'total': sum(years),
            'sample': str(row[10] or '').strip(),
        })
    sources.sort(key=lambda s: -s['total'])

    # ---- รายประเทศ (คำนวณเองจากรายแหล่ง — ชีตในไฟล์เป็นสูตร) ----
    countries = {}
    for s in sources:
        c = countries.setdefault(s['country'], {'country': s['country'], 'centers': 0, 'y': [0, 0, 0, 0], 'total': 0})
        c['centers'] += 1
        for i in range(4):
            c['y'][i] += s['y'][i]
        c['total'] += s['total']
    countries = sorted(countries.values(), key=lambda c: -c['total'])

    # ---- รายเดือน: รวมยอดต่อเดือนจาก Top60 ----
    ws = wb['รายเดือน Top60']
    rows = list(ws.iter_rows(values_only=True))
    import re
    month_labels = []
    for v in rows[0][1:]:
        if v and re.match(r'^[฀-๿.]+6[5-8]$', str(v).strip()):
            month_labels.append(str(v).strip())
        else:
            break  # เจอคอลัมน์รวม/อื่น ๆ หยุดเลย
    n_months = len(month_labels)
    monthly_total = [0] * n_months
    for row in rows[1:]:
        if not row[0]:
            continue
        for i in range(n_months):
            v = row[1 + i]
            monthly_total[i] += int(v) if isinstance(v, (int, float)) else 0

    # ---- quadrant: FB active × ส่งข่าวใน 4 ปี ----
    with open(PROFILES, encoding='utf-8-sig') as f:
        centers = list(csv.DictReader(f))

    def is_active(r):
        return r['active_level'] == 'active'

    def sends(r):
        return int(r['news4y'] or 0) > 0

    quad = {'A': [], 'B': [], 'C': [], 'D': []}
    for r in centers:
        key = ('A' if sends(r) else 'B') if is_active(r) else ('C' if sends(r) else 'D')
        quad[key].append(r)

    def brief(r):
        return {
            'name': r['center'],
            'place': r['province'] or r['country'],
            'followers': '' if r['followers'] in ('', '-') else r['followers'],
            'lastPost': '' if r['last_post'] in ('', '-') else r['last_post'],
            'active': '' if r['active_level'] in ('', '-') else r['active_level'],
            'types': r['activity_types'] if r['activity_types'] not in ('', '-') else '',
            'news4y': int(r['news4y'] or 0),
            'fb': r['facebook_url'],
        }

    def fnum(r):
        s = (r['followers'] or '').upper().replace(',', '')
        try:
            return float(s[:-1]) * 1000 if s.endswith('K') else float(s)
        except ValueError:
            return 0

    quadrant = {
        'counts': {k: len(v) for k, v in quad.items()},
        # กลุ่ม B ทั้งหมด (FB active แต่ไม่มีข่าว 4 ปี) เรียงตาม followers
        'groupB': [brief(r) for r in sorted(quad['B'], key=fnum, reverse=True)],
        # กลุ่ม C ตัวท็อป (ส่งข่าวเยอะแต่หน้าบ้าน FB อ่อน)
        'groupC': [brief(r) for r in sorted(quad['C'], key=lambda r: -int(r['news4y'] or 0))[:12]],
    }

    data = {
        'generated': date.today().isoformat(),
        'source': XLSX.name,
        'headlines': {'total': 8607, 'byYear': [2564, 2151, 2174, 1718], 'identified': 6589},
        'sources': sources,
        'countries': countries,
        'months': {'labels': month_labels, 'total': monthly_total,
                   'note': 'รวมเฉพาะแหล่งข่าว Top 60'},
        'quadrant': quadrant,
    }
    args.out.write_text(json.dumps(data, ensure_ascii=False, separators=(',', ':')), encoding='utf-8')
    print(f'เขียน {args.out}')
    print(f'  แหล่งข่าว {len(sources)} · ประเทศ {len(countries)} · เดือน {n_months}')
    print(f"  quadrant: A={len(quad['A'])} B={len(quad['B'])} C={len(quad['C'])} D={len(quad['D'])}")


if __name__ == '__main__':
    main()
